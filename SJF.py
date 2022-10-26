from openpyxl import load_workbook
workbook = load_workbook(filename="cpu-scheduling.xlsx")
sheet = workbook.active
count=0
for row in sheet.rows:
    count+=1
    #print(row)
#print(count)
    

list_colA=[]
for columnA in sheet.iter_rows(max_row=count,min_row=2,max_col=1,values_only=True):
    list_colA.append(columnA[0])
#print(list_colA)
PID=1
list_colB=[]
for columnB in sheet.iter_rows(max_row=count,min_row=2,max_col=3,min_col=2,values_only=True):
    list_colB.append(columnB[0]) 

list_colC=[]
for columnC in sheet.iter_rows(max_row=count,min_row=2,max_col=4,min_col=3,values_only=True):
    if columnC[0] != None:
        list_colC.append(columnC[0])

list_colD=[]
for columnD in sheet.iter_rows(max_row=count,min_row=2,max_col=5,min_col=4,values_only=True):
    list_colD.append(columnD[0])

allrows=[]
for values in sheet.iter_rows(max_row=count,min_row=2,values_only=True):
    allrows.append(values)

totaltime=count-2
for i in list_colC:
    totaltime+=i
#totale tid å execute er 168

highestinstructionload=max(list_colC)
rowssorted=[]
arrived=[]
stop=0

tempsorted=[]
gtime=1
time=1
maxtmp=0
for nums in range(0,highestinstructionload+1):
    
    for eachrow in allrows:
        if eachrow[2]==nums:
            rowssorted.append(eachrow)
            #satte rowsa i rowssorted i stigende rekkefølge etter kolonneC

def findSJF(time):
    global maxtmp
    global gtime
    global instructionload
    global waiting
    
    temp=[]
    tempsorted=[]
    for i in allrows:
        if i[1]<=time:
            temp.append(i)
            

    for eachrow in temp:
        if eachrow[2] > maxtmp:
            maxtmp=eachrow[2]
    for i in range(0,maxtmp+1):
        for row in temp:
            if row[2]==i:
                tempsorted.append(row)
                #sortert tempsorted etter kolonneC av de som har arrived.
    
#    print("len of tmpsortd",len(tempsorted))   
    if len(tempsorted) >0:
        lentmpsort=len(tempsorted)
        currentPID=tempsorted[0][0]
        instructionload=tempsorted[0][2]-1
        #currentPID er nå av de som arriva, den med lavest instructionload.
        #print(tempsorted)
        #temp=[]
        for i in range(time,time+tempsorted[0][2]+1):
         #   for i in allrows:
          #      if i[1]<=time:
           #         temp.append(i)
            temp=[]
            tempsorted2=[]
            for g in allrows:
                if g[1]<=i:
                    temp.append(g)
            for h in temp:
                if h[0] > maxtmp:
                    maxtmp=h[0]
            for u in range(0,maxtmp+1):
                for y in temp:
                    if y[0]==u:
                        tempsorted2.append(y)
            if len(tempsorted2)==0:
                print("CPU is idle.")
            

            if instructionload != -1:
                print("Time Unit",i,": PID ",currentPID,"executes.",instructionload," instructions left.")
                instructionload -=1
                
                for j in tempsorted2:
                    if j[0] != currentPID:
                        waiting=1+i-int(j[1])
                        #print("PID ",j[0],"wait=",1+int(waiting)-int(j[1]))
                        print("PID ",j[0],"wait= ",waiting)
                if i==totaltime:
                    print("All processes have executed. End of simulation.")
                    break
                    return
            else:
                time+=1
                #if time==totaltime:
                   # print("All processes have executed. End of simulation.")
          
                print("Time Unit",i,": Context Switch.")
                for j in tempsorted2:
                    if j[0] != currentPID:
                        waiting=1+i-int(j[1])
                        print("PID ",j[0],"wait= ",waiting)
        
        time+=tempsorted[0][2]
        gtime=time
        for k in allrows:
            if currentPID==k[0]:
                allrows.remove(k)
    else:
        if time !=totaltime+1:
            print("Time Unit",time,": CPU is idle.")
        
        time+=1        
        gtime=time
for i in range(0,count):
    findSJF(gtime)

    #test